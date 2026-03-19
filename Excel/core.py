"""
core.py — Business logic, WebSocket server (Excel export tool)
Dựa trên Register, nhưng tập trung vào việc export kết quả ra Excel.
"""

import asyncio
import threading
import websockets
import json
import random
import time
import sys
import os
from datetime import datetime

from PyQt6.QtCore import QObject, pyqtSignal

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# ── Shared state ──────────────────────────────────────────────────────────────
sessions   = {}   # {sessionId: {ws, pending, status, connected_at}}
acct_log   = []   # [{session_id, username, email, password, status, note}]

_req_id    = 0
_loop      = None
_gui       = None
_running   = False
_stop_flag = False

ACCOUNT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "account.txt")
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))

# ── Signal bridge (asyncio → Qt) ──────────────────────────────────────────────
class Bridge(QObject):
    log_signal     = pyqtSignal(str)
    refresh_signal = pyqtSignal()

bridge = Bridge()

# ── Helpers ───────────────────────────────────────────────────────────────────
def next_id():
    global _req_id
    _req_id += 1
    return str(_req_id)

def log(msg, _tag="info"):
    now  = datetime.now().strftime("%H:%M:%S")
    line = f"[{now}] {msg}"
    print(line)
    bridge.log_signal.emit(line)

async def human_delay(a=0.4, b=1.2):
    await asyncio.sleep(random.uniform(a, b))

# ── WebSocket communication ───────────────────────────────────────────────────
async def send_and_wait(session_id, action, data, timeout=30):
    s = sessions.get(session_id)
    if not s:
        return None
    req_id = next_id()
    fut    = asyncio.get_event_loop().create_future()
    s["pending"][req_id] = fut
    await s["ws"].send(json.dumps({"requestId": req_id, "action": action, "data": data}))
    log(f"[{str(session_id)[:8]}] >> {action}")
    try:
        return await asyncio.wait_for(asyncio.shield(fut), timeout=timeout)
    except asyncio.TimeoutError:
        log(f"[{str(session_id)[:8]}] !! Timeout: {action}")
        return None
    finally:
        s["pending"].pop(req_id, None)

async def wait_for_selector(session_id, sel, tab_id=None, max_wait=20, interval=0.5):
    data = {"selector": sel}
    if tab_id:
        data["tabId"] = tab_id
    elapsed = 0
    while elapsed < max_wait:
        if _stop_flag:
            return False
        res = await send_and_wait(session_id, "check_element", data, timeout=6)
        if res and (res.get("result") or {}).get("found"):
            return True
        await asyncio.sleep(interval)
        elapsed += interval
    log(f"[{str(session_id)[:8]}] !! Not found: {sel}")
    return False

def set_status(session_id, status, note=""):
    if session_id in sessions:
        sessions[session_id]["status"] = status
    for e in reversed(acct_log):
        if e["session_id"] == session_id:
            e["status"] = status
            if note:
                e["note"] = note
            break
    bridge.refresh_signal.emit()

# ── Load data from Excel ───────────────────────────────────────────────────────
# Mapping: cot Excel → key trong acct_log
EXCEL_COL_MAP = {
    "Ho ten":    "full_name",
    "CCCD":      "cccd",
    "Gioi tinh": "gender",
    "Ngay sinh": "birthday",
    "Dia chi":   "address",
    "Ngay cap":  "issue_date",
    "Han":       "expiry_date",
}

imported_data = []   # list of dict — duoc dien tu file Excel

def load_excel_file(path: str) -> int:
    """
    Doc file Excel theo template CCCD.
    Tra ve so dong doc duoc, luu vao imported_data.
    Header phai khop voi EXCEL_COL_MAP.
    """
    global imported_data
    try:
        import openpyxl
    except ImportError:
        log("!! Thieu openpyxl. Chay: pip install openpyxl")
        return 0

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
    except Exception as ex:
        log(f"!! Khong mo duoc file Excel: {ex}")
        return 0

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log("!! File Excel rong!")
        return 0

    # Doc header row 1
    header_row = [str(h).strip() if h else "" for h in rows[0]]
    col_indices = {}
    for col_key, acct_key in EXCEL_COL_MAP.items():
        try:
            col_indices[acct_key] = header_row.index(col_key)
        except ValueError:
            pass  # cot khong co trong file

    imported_data = []
    for raw in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in raw):
            continue  # bo qua dong trong
        entry = {}
        for acct_key, col_i in col_indices.items():
            val = raw[col_i] if col_i < len(raw) else ""
            entry[acct_key] = str(val).strip() if val is not None else ""
        imported_data.append(entry)

    log(f">> Import Excel: doc duoc {len(imported_data)} dong tu {os.path.basename(path)}")
    bridge.refresh_signal.emit()
    return len(imported_data)


# ── Automation ────────────────────────────────────────────────────────────────
TARGET_URL = (
    "https://m.luxshare-ict.com/hr/idcardcollectforvnintroducer.html"
    "?introducer=DUCMINH-202507&zarsrc=1303"
    "&utm_source=zalo&utm_medium=zalo&utm_campaign=zalo"
)

async def run_task(session_id, username, email, password):
    """
    Mo trang Luxshare CCCD, dien form va submit.
    username = full_name (lay tu acct_log theo session_id)
    """
    sid = str(session_id)[:8]

    # Lay du lieu CCCD cho session nay
    row_data = {}
    for e in acct_log:
        if e["session_id"] == session_id:
            row_data = e
            break

    full_name   = row_data.get("full_name",   "")
    cccd        = row_data.get("cccd",        "")
    gender      = row_data.get("gender",      "")   # "Nam" hoac "Nu"
    birthday    = row_data.get("birthday",    "")   # dd/mm/yyyy
    address     = row_data.get("address",     "")
    issue_date  = row_data.get("issue_date",  "")
    expiry_date = row_data.get("expiry_date", "")

    log(f"[{sid}] Bat dau: {full_name} — {cccd}")
    set_status(session_id, "Dang chay")

    try:
        # ── 1. Mo tab moi vao trang Luxshare ──────────────────────────────────
        res    = await send_and_wait(session_id, "open_url",
                                     {"url": TARGET_URL, "newTab": True}, timeout=15)
        tab_id = (res or {}).get("result", {}).get("tabId")

        def tab(x={}):
            d = dict(x)
            if tab_id:
                d["tabId"] = tab_id
            return d

        async def sw(action, data, **kw):
            return await send_and_wait(session_id, action, tab(data), **kw)

        async def wfs(sel, max_wait=20):
            return await wait_for_selector(session_id, sel, tab_id, max_wait=max_wait)

        # Ham fill: clear roi go tung ky tu (nhu Register)
        async def fill(sel: str, value: str) -> bool:
            try:
                await send_and_wait(session_id, "clear_field",
                                    tab({"selector": sel}), timeout=8)
                await send_and_wait(session_id, "type_text",
                                    tab({"selector": sel, "value": value}), timeout=30)
                return True
            except Exception as ex:
                log(f"[{sid}] fill ERR [{sel[:40]}]: {ex}")
                return False



        log(f"[{sid}] Tai trang — cho 5s...")
        await asyncio.sleep(5)

        # ── 2. Ho va ten — wfs chac chan element co trong DOM truoc khi fill ────
        if await wfs('input[data-input-clear="1"]', max_wait=20):
            ok = await fill('input[data-input-clear="1"]', full_name)
            log(f"[{sid}] Ho ten: {full_name} — {'OK' if ok else 'FAIL'}")
        else:
            log(f"[{sid}] !! Khong tim thay field Ho ten — trang chua load?")

        # ── 3. So CCCD ────────────────────────────────────────────────────────
        if await wfs('input[data-input-clear="2"]', max_wait=10):
            ok = await fill('input[data-input-clear="2"]', cccd)
            log(f"[{sid}] CCCD: {cccd} — {'OK' if ok else 'FAIL'}")

        # ── 4. Gioi tinh (readonly) ───────────────────────────────────────────
        if gender and await wfs("input.gender", max_wait=10):
            ok = await fill("input.gender", gender)
            log(f"[{sid}] Gioi tinh: {gender} — {'OK' if ok else 'FAIL'}")

        # ── 5. Ngay sinh ──────────────────────────────────────────────────────
        if birthday and await wfs('input[data-field="birthday"]', max_wait=10):
            ok = await fill('input[data-field="birthday"]', _to_iso(birthday))
            log(f"[{sid}] Ngay sinh: {birthday} → {_to_iso(birthday)} — {'OK' if ok else 'FAIL'}")

        # ── 6. Noi thuong tru ─────────────────────────────────────────────────
        if address and await wfs('input[data-input-clear="5"]', max_wait=10):
            ok = await fill('input[data-input-clear="5"]', address)
            log(f"[{sid}] Dia chi: {address[:50]} — {'OK' if ok else 'FAIL'}")

        # ── 7. Ngay cap ───────────────────────────────────────────────────────
        if issue_date and await wfs('input[data-field="efectiveStartDate"]', max_wait=10):
            ok = await fill('input[data-field="efectiveStartDate"]', _to_iso(issue_date))
            log(f"[{sid}] Ngay cap: {issue_date} — {'OK' if ok else 'FAIL'}")

        # ── 8. Han CCCD ───────────────────────────────────────────────────────
        if expiry_date and await wfs('input[data-field="efectiveEndDate"]', max_wait=10):
            ok = await fill('input[data-field="efectiveEndDate"]', _to_iso(expiry_date))
            log(f"[{sid}] Han: {expiry_date} — {'OK' if ok else 'FAIL'}")

        await asyncio.sleep(0.5)

        # ── 9. CAPTCHA: lay anh base64 → luu file PNG ─────────────────────────
        SEL_CAP_IMG = 'img[src^="data:image"]'
        SEL_CAP_INP = 'input[data-input-clear="8"]'

        res_img = await send_and_wait(session_id, "execute_script",
                                       tab({"script":
                                            "var i=document.querySelector('img[src^=\"data:image\"]');i?i.src:''"}),
                                       timeout=8)
        cap_src = str((res_img or {}).get("result", ""))
        cap_val = ""
        if cap_src:
            log(f"[{sid}] Co captcha (base64 {len(cap_src)} ky tu) — luu file PNG")
            try:
                import base64 as _b64, re as _re
                m = _re.match(r"data:image/\w+;base64,(.+)", cap_src)
                if m:
                    cap_path = os.path.join(BASE_DIR, f"captcha_{sid}.png")
                    with open(cap_path, "wb") as f:
                        f.write(_b64.b64decode(m.group(1)))
                    log(f"[{sid}] Captcha luu: {cap_path}")
            except Exception as ce:
                log(f"[{sid}] Luu captcha loi: {ce}")
        else:
            log(f"[{sid}] Khong co captcha")

        if cap_val:
            await fill(SEL_CAP_INP, cap_val)

        await asyncio.sleep(0.3)


        # ── 10. Cap nhat acct_log va hoan thanh ────────────────────────────────
        for e in reversed(acct_log):
            if e["session_id"] == session_id:
                e["note"] = "Da dien form" + (" (co captcha)" if cap_src else "")
                break

        set_status(session_id, "Cho captcha" if cap_src else "Hoan thanh",
                   "Co captcha — can xu ly" if cap_src else "Dien xong")
        log(f"[{sid}] Hoan thanh dien form: {full_name}")

    except Exception as ex:
        set_status(session_id, "Loi", str(ex)[:80])
        log(f"[{sid}] Error: {ex}")


def _to_iso(date_str: str) -> str:
    """Chuyen dd/mm/yyyy -> yyyy-mm-dd (input[type=date])."""
    date_str = date_str.strip()
    parts = date_str.replace("-", "/").split("/")
    if len(parts) == 3:
        d, m, y = parts[0], parts[1], parts[2]
        # Neu da la yyyy-mm-dd thi giu nguyen
        if len(y) == 2:
            y = "20" + y
        if len(d) == 4:   # dau vao la yyyy/mm/dd
            return f"{d}-{m}-{y}"
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return date_str  # tra nguyen neu khong parse duoc






async def run_all_sessions():
    global _running, _stop_flag
    _stop_flag = False

    if not imported_data:
        log("!! Chua co du lieu — bam 'Import Excel' truoc!")
        _running = False
        bridge.refresh_signal.emit()
        return

    session_ids = list(sessions.keys())
    if not session_ids:
        log("!! Chua co session nao ket noi!")
        _running = False
        bridge.refresh_signal.emit()
        return

    log(f">> Bat dau {len(session_ids)} session(s) voi {len(imported_data)} dong du lieu...")
    tasks = []
    for sid, row in zip(session_ids, imported_data):
        entry = {
            "session_id":  sid,
            "full_name":   row.get("full_name",  ""),
            "cccd":        row.get("cccd",        ""),
            "gender":      row.get("gender",      ""),
            "birthday":    row.get("birthday",    ""),
            "address":     row.get("address",     ""),
            "issue_date":  row.get("issue_date",  ""),
            "expiry_date": row.get("expiry_date", ""),
            "status":      "Cho...",
            "note":        "",
            "cookie":      "",
        }
        acct_log.append(entry)
        log(f"  >> [{str(sid)[:8]}] {row.get('full_name','?')} — {row.get('cccd','?')}")
        tasks.append(asyncio.create_task(
            run_task(sid, row.get("full_name", ""), "", "")
        ))

    bridge.refresh_signal.emit()
    await asyncio.gather(*tasks, return_exceptions=True)
    log(">> Tat ca sessions hoan thanh!")
    _running = False
    bridge.refresh_signal.emit()



# ── WebSocket server ───────────────────────────────────────────────────────────
async def ws_handler(websocket):
    session_id = None
    try:
        async for raw in websocket:
            try:
                msg = json.loads(raw)
            except Exception as ex:
                log(f"Parse error: {ex}")
                continue
            t = msg.get("type")
            if t == "register":
                session_id = msg.get("sessionId")
                if not session_id:
                    continue
                sessions[session_id] = {
                    "ws": websocket, "info": msg,
                    "pending": {}, "status": "Ket noi",
                    "connected_at": time.time()
                }
                log(f"++ Session: {str(session_id)[:8]}... ({len(sessions)} total)")
                bridge.refresh_signal.emit()
            elif t == "result":
                sid = msg.get("sessionId") or session_id
                if sid and sid in sessions:
                    rid = msg.get("requestId")
                    p   = sessions[sid]["pending"]
                    if rid and rid in p and not p[rid].done():
                        p[rid].set_result(msg)
    except websockets.exceptions.ConnectionClosed:
        pass
    except Exception as ex:
        log(f"Handler error: {ex}")
    finally:
        if session_id and session_id in sessions:
            sessions.pop(session_id, None)
            log(f"-- Session ngat: {str(session_id)[:8]}... ({len(sessions)} con)")
            bridge.refresh_signal.emit()

async def _ws_server():
    await websockets.serve(ws_handler, "127.0.0.1", 8000)
    log("[WS] Server san sang: ws://127.0.0.1:8000")
    await asyncio.Future()

def start_asyncio_loop():
    global _loop
    _loop = asyncio.new_event_loop()
    asyncio.set_event_loop(_loop)
    _loop.run_until_complete(_ws_server())

# ── Public triggers ────────────────────────────────────────────────────────────
def trigger_run():
    global _running
    if _running:
        return
    _running = True
    asyncio.run_coroutine_threadsafe(run_all_sessions(), _loop)

def trigger_stop():
    global _stop_flag, _running
    _stop_flag = True
    _running   = False
    log("!! Dung — cac session se ket thuc sau buoc hien tai")
    bridge.refresh_signal.emit()

# ── Export Excel ───────────────────────────────────────────────────────────────
# Mapping header hien thi  →  key trong acct_log
EXCEL_FIELDS = [
    ("#",          None),
    ("Username",   "username"),
    ("Ho ten",     "full_name"),
    ("CCCD",       "cccd"),
    ("Gioi tinh",  "gender"),
    ("Ngay sinh",  "birthday"),
    ("Dia chi",    "address"),
    ("Ngay cap",   "issue_date"),
    ("Han",        "expiry_date"),
    ("Status",     "status"),
    ("Ghi chu",    "note"),
    ("Cookie",     "cookie"),
]
COL_WIDTHS = [5, 18, 22, 14, 10, 12, 30, 12, 12, 12, 20, 60]

def export_excel(path: str) -> int:
    """Xuat acct_log ra file Excel .xlsx voi cac truong CCCD."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log("!! Thieu thu vien openpyxl. Chay: pip install openpyxl")
        return 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Du lieu CCCD"

    header_font  = Font(bold=True, color="FFFFFF", name="Segoe UI", size=11)
    header_fill  = PatternFill("solid", fgColor="1A3A5C")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    STATUS_FILLS = {
        "Hoan thanh": "1A4A2A",
        "Loi":        "4A1A1A",
        "Dang chay":  "1A2A4A",
        "Cho...":     "2A2A2A",
    }

    # Header row
    for col_i, ((hdr, _), w) in enumerate(zip(EXCEL_FIELDS, COL_WIDTHS), 1):
        c = ws.cell(row=1, column=col_i, value=hdr)
        c.font = header_font; c.fill = header_fill
        c.alignment = center_align; c.border = thin
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    # Data rows
    CENTER_KEYS = {"status", "gender", "birthday", "issue_date", "expiry_date", "cccd"}
    for row_i, e in enumerate(acct_log, 2):
        status   = e.get("status", "")
        fill_hex = STATUS_FILLS.get(status, "1C2333")
        rfill    = PatternFill("solid", fgColor=fill_hex)
        rfont    = Font(name="Segoe UI", size=10, color="D0DEFA")

        for col_i, (hdr, key) in enumerate(EXCEL_FIELDS, 1):
            val = (row_i - 1) if key is None else e.get(key, "")
            aln = center_align if (key is None or key in CENTER_KEYS) else left_align
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill = rfill; c.alignment = aln; c.border = thin; c.font = rfont

        ws.row_dimensions[row_i].height = 18

    ws.freeze_panes = "A2"
    wb.save(path)
    log(f">> Da xuat {len(acct_log)} dong ra: {path}")
    return len(acct_log)
