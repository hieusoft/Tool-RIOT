"""
core_register.py — Business logic cho chức năng Đăng ký tài khoản (không có WS server)
WS server được quản lý tập trung bởi gui.py
"""

import asyncio
import json
import random
import time
import sys
import re
import os
import urllib.parse
from datetime import datetime
from faker import Faker

from PyQt6.QtCore import QObject, pyqtSignal

fake = Faker()

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# ── Shared state ──────────────────────────────────────────────────────────────
sessions   = {}   # {sessionId: {ws, pending, status, connected_at}}
acct_log   = []   # [{session_id, username, email, password, status}]

_req_id    = 0
_loop      = None
_running   = False
_stop_flag = False

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
ACCOUNT_FILE = os.path.join(BASE_DIR, "account_register.xlsx")

# ── Signal bridge (asyncio → Qt) ──────────────────────────────────────────────
class Bridge(QObject):
    log_signal     = pyqtSignal(str)
    refresh_signal = pyqtSignal()

bridge = Bridge()

# ── Helpers ───────────────────────────────────────────────────────────────────
def next_id():
    global _req_id
    _req_id += 1
    return str(_req_id)

def log(msg, _tag="info"):
    now  = datetime.now().strftime("%H:%M:%S")
    line = f"[{now}] {msg}"
    print(line)
    bridge.log_signal.emit(line)

async def human_delay(a=0.4, b=1.2):
    await asyncio.sleep(random.uniform(a, b))

# ── WebSocket communication ───────────────────────────────────────────────────
async def send_and_wait(session_id, action, data, timeout=30):
    s = sessions.get(session_id)
    if not s:
        return None
    req_id = next_id()
    fut    = asyncio.get_event_loop().create_future()
    s["pending"][req_id] = fut
    await s["ws"].send(json.dumps({"requestId": req_id, "action": action, "data": data}))
    log(f"[{str(session_id)[:8]}] >> {action}")
    try:
        return await asyncio.wait_for(asyncio.shield(fut), timeout=timeout)
    except asyncio.TimeoutError:
        log(f"[{str(session_id)[:8]}] !! Timeout: {action}")
        return None
    finally:
        s["pending"].pop(req_id, None)

async def wait_for_selector(session_id, sel, tab_id=None, max_wait=20, interval=0.4):
    data = {"selector": sel}
    if tab_id:
        data["tabId"] = tab_id
    elapsed = 0
    while elapsed < max_wait:
        if _stop_flag:
            return False
        res = await send_and_wait(session_id, "check_element", data, timeout=5)
        if res and (res.get("result") or {}).get("found"):
            return True
        await asyncio.sleep(interval)
        elapsed += interval
    log(f"[{str(session_id)[:8]}] !! Not found: {sel}")
    return False

def set_status(session_id, status):
    if session_id in sessions:
        sessions[session_id]["status"] = status
    for e in reversed(acct_log):
        if e["session_id"] == session_id:
            e["status"] = status
            break
    bridge.refresh_signal.emit()

# ── Load accounts ─────────────────────────────────────────────────────────────
def load_accounts():
    """
    Đọc file account_register.xlsx
    Sheet1: cột A=email, B=username, C=password
    """
    accounts = []
    if not os.path.exists(ACCOUNT_FILE):
        log(f"!! Khong tim thay file: {ACCOUNT_FILE}")
        return accounts
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ACCOUNT_FILE, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            if not row or not row[0]:
                continue
            email    = str(row[0]).strip()
            username = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            password = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if email:
                accounts.append({"email": email, "username": username, "password": password})
        wb.close()
    except Exception as e:
        log(f"!! Loi doc file Excel: {e}")
    log(f">> Doc duoc {len(accounts)} tai khoan tu {os.path.basename(ACCOUNT_FILE)}")
    return accounts

# ── Automation ────────────────────────────────────────────────────────────────
async def run_signup(session_id, email, username, password):
    sid = str(session_id)[:8]

    if _stop_flag:
        set_status(session_id, "Da dung")
        return
    log(f"[{sid}] Start: {email} / {username}")
    set_status(session_id, "Dang chay")

    sw  = lambda a, d, **kw: send_and_wait(session_id, a, d, **kw)
    wfs = lambda s, **kw: wait_for_selector(session_id, s, **kw)

    try:
        SIGNUP_URL = "https://signup.leagueoflegends.com/en-us/signup/index#/"
        await sw("open_url", {"url": SIGNUP_URL, "newTab": True})

        if await wfs('[data-testid="riot-signup-email"]'):
            await human_delay(0.5, 1.0)
            await sw("type_text", {"selector": '[data-testid="riot-signup-email"]', "value": email}, timeout=30)

        await human_delay(0.6, 1.4)
        await sw("click", {"selector": "#newsletter"})
        await human_delay(0.3, 0.8)
        await sw("click", {"selector": "#thirdpartycomms"})
        await human_delay(0.8, 1.8)
        await sw("click", {"selector": '[data-testid="btn-signup-submit"]'})

        if await wfs('[data-testid="riot-signup-username"]'):
            await human_delay()
            await sw("type_text", {"selector": '[data-testid="riot-signup-username"]', "value": username}, timeout=20)

        await human_delay(0.7, 1.5)
        await sw("click", {"selector": '[data-testid="btn-signup-submit"]'})

        if await wfs('[data-testid="input-password"]'):
            await human_delay()
            await sw("type_text", {"selector": '[data-testid="input-password"]', "value": password}, timeout=20)

        if await wfs('[data-testid="password-confirm"]'):
            await human_delay(0.4, 0.9)
            await sw("type_text", {"selector": '[data-testid="password-confirm"]', "value": password}, timeout=20)

        await human_delay(0.7, 1.5)
        await sw("click", {"selector": '[data-testid="btn-signup-submit"]'})

        if await wfs('#tos-scrollable-area', max_wait=20):
            await human_delay()
            for pos in [500, 1500, 3000, 999999]:
                await sw("scroll_element", {"selector": "#tos-scrollable-area", "top": pos})
                await human_delay(0.4, 0.9)
            if await wfs('#tos-checkbox:not([disabled])', max_wait=10):
                await human_delay(0.3, 0.7)
                await sw("click", {"selector": "#tos-checkbox"})
                await human_delay(0.4, 0.8)
                if await wfs('[data-testid="btn-accept-tos"]:not([disabled])', max_wait=10):
                    await human_delay(0.5, 1.0)
                    await sw("click", {"selector": '[data-testid="btn-accept-tos"]'})
                    await human_delay(0.7, 1.5)
                    await human_delay(0.7, 1.5)

                    # Kiểm tra lỗi "Username must be unique"
                    err_sel = '.errorMessage'
                    while True:
                        err_res   = await send_and_wait(session_id, "check_element",
                                                        {"selector": err_sel}, timeout=5)
                        err_found = (err_res or {}).get("result", {}).get("found", False)
                        if not err_found:
                            break
                        txt_res  = await send_and_wait(session_id, "get_text",
                                                       {"selector": err_sel}, timeout=5)
                        err_text = str((txt_res or {}).get("result", {}).get("text", ""))
                        if "unique" not in err_text.lower() and "username" not in err_text.lower():
                            break
                        log(f"[{sid}] Username trung lap — doi username moi...")
                        base     = re.sub(r'[^a-zA-Z0-9_]', '', fake.user_name())[:13] or "user"
                        username = base + fake.numerify("##")
                        await sw("clear_text", {"selector": '[data-testid="riot-signup-username"]'})
                        await human_delay(0.3, 0.6)
                        await sw("type_text", {"selector": '[data-testid="riot-signup-username"]',
                                               "value": username}, timeout=20)
                        await human_delay(0.5, 1.0)
                        await sw("click", {"selector": '[data-testid="btn-signup-submit"]'})
                        await human_delay(1.0, 2.0)

                    captcha_sel = 'iframe[src*="hcaptcha.com"]'
                    log(f"[{sid}] Kiem tra hCaptcha...")
                    while True:
                        res   = await send_and_wait(session_id, "check_element",
                                                    {"selector": captcha_sel}, timeout=5)
                        found = (res or {}).get("result", {}).get("found", False)
                        if not found:
                            break
                        log(f"[{sid}] hCaptcha dang hien — cho 5s...")
                        await asyncio.sleep(5)
                    log(f"[{sid}] hCaptcha da bien mat, tiep tuc...")

        # ── Lấy token từ active tab (sau redirect về localhost) ──
        set_status(session_id, "Lay token...")
        log(f"[{sid}] Dang lay token...")

        GET_TOKEN  = "https://auth.riotgames.com/authorize?redirect_uri=http://localhost/redirect&client_id=riot-client&response_type=token%20id_token&nonce=1&scope=openid%20link%20ban%20lol_region%20account"
        LOGOUT_URL = "https://login.riotgames.com/end-session-redirect?redirect_uri=https%3A%2F%2Fauth.riotgames.com%2Flogout"

        await sw("open_url", {"url": GET_TOKEN})
        await asyncio.sleep(2)

        access_token = ""
        token_tab_id = None
        for i in range(40):
            await asyncio.sleep(0.5)
            tabs_res = await send_and_wait(session_id, "list_tabs", {}, timeout=5)
            tabs     = []
            if tabs_res:
                r = tabs_res.get("result", [])
                tabs = r if isinstance(r, list) else []
            if i % 6 == 0:
                log(f"[{sid}] [DBG] so tab={len(tabs)}")
            for t in tabs:
                url = t.get("url", "") if isinstance(t, dict) else ""
                if "access_token=" in url:
                    frag   = url.split("#", 1)[-1] if "#" in url else url.split("?", 1)[-1]
                    params = dict(urllib.parse.parse_qsl(frag))
                    tok    = params.get("access_token", "")
                    if tok:
                        access_token = tok
                        token_tab_id = t.get("id") or t.get("tabId")
                        break
            if access_token:
                break
        if token_tab_id:
            await send_and_wait(session_id, "close_tab", {"tabId": token_tab_id}, timeout=5)

        if access_token:
            log(f"[{sid}] Token: {access_token[:40]}...")
            for e in reversed(acct_log):
                if e["session_id"] == session_id:
                    e["token"] = access_token
                    break
            bridge.refresh_signal.emit()
        else:
            log(f"[{sid}] Khong lay duoc token!")
            for e in reversed(acct_log):
                if e["session_id"] == session_id:
                    e["token"] = ""
                    break

        set_status(session_id, "Hoan thanh")
        log(f"[{sid}] Logout va don dep tab...")

        # Logout + clear cookie
        await sw("open_url", {"url": LOGOUT_URL})
        await sw("clear_cookies", {})
        await asyncio.sleep(2)

        # Đóng tất cả tab trừ 1
        log(f"[{sid}] Don dep tab cu (giu lai 1)...")
        all_tabs_res = await send_and_wait(session_id, "list_tabs", {}, timeout=5)
        all_tabs = ((all_tabs_res or {}).get("result", [])) or []
        for t in all_tabs[1:]:
            tid = t.get("id") if isinstance(t, dict) else None
            if tid:
                await send_and_wait(session_id, "close_tab", {"tabId": tid}, timeout=5)
        await asyncio.sleep(1)

    except Exception as e:
        set_status(session_id, "Loi")
        log(f"[{sid}] Error: {e}")
        await asyncio.sleep(3)



async def run_all_sessions():
    global _running, _stop_flag
    _stop_flag = False

    accounts = load_accounts()
    if not accounts:
        log("!! Khong co tai khoan nao trong account_register.xlsx (dinh dang: email | username | password)")
        _running = False
        bridge.refresh_signal.emit()
        return

    session_ids = list(sessions.keys())
    if not session_ids:
        log("!! Chua co session nao ket noi!")
        _running = False
        bridge.refresh_signal.emit()
        return

    log(f">> Bat dau {len(session_ids)} session(s) voi {len(accounts)} tai khoan...")

    queue: asyncio.Queue = asyncio.Queue()
    for acct in accounts:
        await queue.put(acct)

    async def session_worker(sid):
        while not _stop_flag:
            try:
                acct = queue.get_nowait()
            except asyncio.QueueEmpty:
                break
            email    = acct["email"]
            username = acct["username"]
            password = acct["password"]
            entry = {
                "session_id": sid, "username": username,
                "email": email, "password": password, "status": "Cho...",
            }
            acct_log.append(entry)
            bridge.refresh_signal.emit()
            log(f"  >> [{str(sid)[:8]}] {email} / {username}")
            await run_signup(sid, email, username, password)
            queue.task_done()

    tasks = [asyncio.create_task(session_worker(sid)) for sid in session_ids]
    bridge.refresh_signal.emit()
    await asyncio.gather(*tasks, return_exceptions=True)
    log(">> Tat ca sessions hoan thanh!")
    _running = False
    bridge.refresh_signal.emit()

# ── Public triggers (called from GUI) ─────────────────────────────────────────
def trigger_run(loop):
    global _running
    if _running:
        return
    _running = True
    asyncio.run_coroutine_threadsafe(run_all_sessions(), loop)

def trigger_stop():
    global _stop_flag, _running
    _stop_flag = True
    _running   = False
    log("!! Dung — cac session se ket thuc sau buoc hien tai")
    bridge.refresh_signal.emit()
