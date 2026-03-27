"""
gui.py — App chính Tool-RIOT: QTabWidget với 3 tab (Login, Change Password, Register)
WS server chạy tập trung, route session đến từng module core tương ứng.
"""

import time
import os
import asyncio
import websockets
import json
import threading

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem,
    QListWidget, QListWidgetItem, QTextEdit, QTabWidget,
    QHeaderView, QAbstractItemView, QMessageBox, QFileDialog
)
from PyQt6.QtCore import Qt, QTimer, QObject, pyqtSignal
from PyQt6.QtGui import QColor, QBrush

import core_login
import core_changepass
import core_register

# ── Global asyncio loop (shared) ─────────────────────────────────────────────
_loop = None

# session_id → module name ("login" | "changepass" | "register")
_session_module_map = {}

# ── Stylesheet chung ──────────────────────────────────────────────────────────
STYLE = """
QMainWindow, QWidget#root {
    background: #0c0f14;
}
QWidget {
    background: transparent;
    color: #eef2f6;
    font-family: "Segoe UI", "Inter", sans-serif;
    font-size: 13px;
}
QTabWidget::pane {
    border: none;
    background: #0c0f14;
}
QTabBar::tab {
    background: #101520;
    color: #7a9abf;
    padding: 10px 28px;
    border: none;
    border-bottom: 3px solid transparent;
    font-size: 13px;
    font-weight: bold;
    min-width: 100px;
}
QTabBar::tab:selected {
    background: #131b28;
    color: #ffffff;
    border-bottom: 3px solid #2eb8c8;
}
QTabBar::tab:hover:!selected {
    background: #141c2a;
    color: #bdd2ff;
}
QWidget#toolbar {
    background: #0f131b;
    border-bottom: 2px solid #262d38;
}
QWidget#left_panel {
    background: #0e121a;
    border-right: 1px solid #262d38;
}
QWidget#left_header {
    background: #121a24;
    border-bottom: 2px solid #263040;
}
QListWidget {
    background: #0e121a;
    border: none;
    outline: none;
}
QListWidget::item {
    border-bottom: 1px solid #202734;
    padding: 10px 12px;
    color: #d0defa;
}
QListWidget::item:selected {
    background: #182337;
    border-left: 3px solid #2eb8c8;
    color: #ffffff;
}
QListWidget::item:hover { background: #17202d; }
QTableWidget {
    background: #0f131b;
    alternate-background-color: #111620;
    gridline-color: #1f2630;
    border: none;
    outline: none;
}
QTableWidget::item {
    padding: 6px 10px;
    border: none;
}
QTableWidget::item:selected { background: #1a2848; color: #ffffff; }
QHeaderView::section {
    background: #10161f;
    color: #d3e2ff;
    font-weight: bold;
    padding: 10px;
    border: none;
    border-bottom: 2px solid #263040;
    border-right: 1px solid #232b36;
}
QPushButton#btn_start_login {
    background: #0e4d6b;
    color: white;
    border: 1px solid #1a88b8;
    padding: 7px 20px;
    font-weight: bold;
    font-size: 13px;
}
QPushButton#btn_start_login:hover    { background: #126090; }
QPushButton#btn_start_login:disabled { background: #1a2a3a; color: #556; }
QPushButton#btn_start_changepass {
    background: #7b2d00;
    color: white;
    border: 1px solid #c44a00;
    padding: 7px 20px;
    font-weight: bold;
    font-size: 13px;
}
QPushButton#btn_start_changepass:hover    { background: #9e3900; }
QPushButton#btn_start_changepass:disabled { background: #2a1a1a; color: #556; }
QPushButton#btn_start_register {
    background: #0e6b3e;
    color: white;
    border: 1px solid #1d9e5d;
    padding: 7px 20px;
    font-weight: bold;
    font-size: 13px;
}
QPushButton#btn_start_register:hover    { background: #12864e; }
QPushButton#btn_start_register:disabled { background: #1a3a2a; color: #556; }
QPushButton#btn_stop {
    background: #8b2a2a;
    color: white;
    border: 1px solid #c23b3b;
    padding: 7px 20px;
    font-weight: bold;
    font-size: 13px;
}
QPushButton#btn_stop:hover    { background: #a33; }
QPushButton#btn_stop:disabled { background: #2a1a1a; color: #556; }
QPushButton#btn_clear {
    background: #1b2331;
    color: #8da2cc;
    border: 1px solid #2f3c51;
    padding: 4px 12px;
    font-size: 11px;
}
QPushButton#btn_clear:hover { background: #232e42; }
QPushButton#btn_export {
    background: #1a3050;
    color: #7ab0ff;
    border: 1px solid #2f5080;
    padding: 7px 16px;
    font-weight: bold;
    font-size: 13px;
}
QPushButton#btn_export:hover { background: #1f3d66; }
QTextEdit#log_box {
    background: #090d12;
    color: #7a9cc0;
    border: none;
    font-family: "Consolas", monospace;
    font-size: 11px;
}
QScrollBar:vertical {
    background: #0b0f15;
    width: 6px;
    border: none;
}
QScrollBar::handle:vertical {
    background: #2f405b;
    border-radius: 3px;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QLabel#status_bar {
    background: #080b0f;
    color: #4a6080;
    padding: 4px 14px;
    font-size: 11px;
    border-top: 1px solid #1a2030;
}
QLabel#count_badge {
    background: #1e2a3a;
    color: #bdd2ff;
    font-size: 11px;
    font-weight: bold;
    padding: 2px 10px;
    border: 1px solid #3d506e;
}
"""

# ── Helper: stat card ─────────────────────────────────────────────────────────
def make_stat_card(parent_layout, label, value, color, last=False):
    card = QWidget()
    card.setObjectName("stat_card")
    border = "border-right:1px solid #262d38;" if not last else ""
    card.setStyleSheet(f"QWidget#stat_card {{ background:#0f131b; {border} }}")
    v = QVBoxLayout(card)
    v.setContentsMargins(20, 14, 20, 14)
    lbl = QLabel(label)
    lbl.setStyleSheet("color:#7a9abf; font-size:10px; font-weight:bold; background:transparent;")
    v.addWidget(lbl)
    num = QLabel(value)
    num.setStyleSheet(f"color:{color}; font-size:26px; font-weight:bold; background:transparent;")
    v.addWidget(num)
    parent_layout.addWidget(card)
    return num

STATUS_COLORS = {
    "Hoan thanh":      ("#112a1c", "#4ae98c"),
    "Co cookie":       ("#0a2010", "#30ff80"),
    "Khong co cookie": ("#1a2a1a", "#80c080"),
    "Lay cookie...":   ("#0e1a2a", "#7ab0ff"),
    "Lay token...":    ("#0e1a2a", "#c0a0ff"),
    "Loi":             ("#2a1111", "#ff7070"),
    "Login loi":       ("#2a1111", "#ff7070"),
    "Da dung":         ("#1a1a1a", "#888888"),
    "Dang chay":       ("#0e1a2a", "#7ab0ff"),
    "Can 2FA":         ("#2a1e0a", "#e8a030"),
    "Can kiem tra":    ("#1a1a2a", "#a090d0"),
    "Cho...":          ("#111620", "#5a7090"),
    "Ket noi":         ("#111a24", "#4a80b0"),
    "Khong co token":  ("#1a1a20", "#9090b0"),
}

def make_cell(text, bg, fg, align=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter):
    it = QTableWidgetItem(str(text))
    it.setBackground(QBrush(bg))
    it.setForeground(QBrush(fg))
    it.setTextAlignment(align)
    return it

CENTER = Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter


# ══════════════════════════════════════════════════════════════════════════════
# Tab 1: LOGIN
# ══════════════════════════════════════════════════════════════════════════════
class LoginTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build()
        core_login.bridge.log_signal.connect(self._on_log)
        core_login.bridge.refresh_signal.connect(self._on_refresh)

        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(1000)

    def _build(self):
        root_v = QVBoxLayout(self)
        root_v.setContentsMargins(0, 0, 0, 0)
        root_v.setSpacing(0)

        # ── Toolbar ──
        tb = QWidget(); tb.setObjectName("toolbar"); tb.setFixedHeight(50)
        h = QHBoxLayout(tb); h.setContentsMargins(16, 0, 16, 0); h.setSpacing(8)
        t = QLabel("Dang nhap — Riot Support")
        t.setStyleSheet("color:#8da2cc; font-size:12px; font-weight:bold; background:transparent;")
        h.addWidget(t)
        h.addStretch()

        self.btn_stop  = QPushButton("■  STOP")
        self.btn_stop.setObjectName("btn_stop"); self.btn_stop.setFixedHeight(32)
        self.btn_stop.setEnabled(False); self.btn_stop.clicked.connect(core_login.trigger_stop)

        self.btn_export = QPushButton("XUAT EXCEL")
        self.btn_export.setObjectName("btn_export"); self.btn_export.setFixedHeight(32)
        self.btn_export.clicked.connect(self._export_excel)

        self.btn_start = QPushButton("▶  BAT DAU")
        self.btn_start.setObjectName("btn_start_login"); self.btn_start.setFixedHeight(32)
        self.btn_start.clicked.connect(self._on_start)

        for w in [self.btn_stop, self.btn_export, self.btn_start]:
            h.addWidget(w)
        root_v.addWidget(tb)

        # ── Body ──
        body = QHBoxLayout()
        body.setContentsMargins(0, 0, 0, 0); body.setSpacing(0)

        # Left
        left = QWidget(); left.setObjectName("left_panel"); left.setFixedWidth(230)
        lv = QVBoxLayout(left); lv.setContentsMargins(0,0,0,0); lv.setSpacing(0)
        lhdr = QWidget(); lhdr.setObjectName("left_header"); lhdr.setFixedHeight(46)
        lh = QHBoxLayout(lhdr); lh.setContentsMargins(12, 0, 12, 0)
        lh.addWidget(QLabel("Session ID", styleSheet="color:white;font-weight:bold;background:transparent;"))
        lh.addStretch()
        self.count_badge = QLabel("0 ket noi"); self.count_badge.setObjectName("count_badge")
        lh.addWidget(self.count_badge)
        lv.addWidget(lhdr)
        self.sess_list = QListWidget(); lv.addWidget(self.sess_list, 1)
        body.addWidget(left)

        # Right
        right_w = QWidget(); right_v = QVBoxLayout(right_w)
        right_v.setContentsMargins(0,0,0,0); right_v.setSpacing(0)

        cards_w = QWidget()
        cards_w.setStyleSheet("background:#0b0e14; border-bottom:1px solid #262d38;")
        cards_h = QHBoxLayout(cards_w); cards_h.setContentsMargins(0,0,0,0); cards_h.setSpacing(0)
        self.s_conn    = make_stat_card(cards_h, "KET NOI",    "0", "#4282ff", last=False)
        self.s_running = make_stat_card(cards_h, "DANG CHAY",  "0", "#fdb45a", last=False)
        self.s_done    = make_stat_card(cards_h, "HOAN THANH", "0", "#4ae98c", last=False)
        self.s_mfa     = make_stat_card(cards_h, "CAN 2FA",    "0", "#e8a030", last=False)
        self.s_error   = make_stat_card(cards_h, "LOI",        "0", "#ff7070", last=True)
        cards_w.setFixedHeight(90); right_v.addWidget(cards_w)

        sec = QWidget(); sec.setFixedHeight(34)
        sec.setStyleSheet("background:#121a24; border-bottom:1px solid #263040;")
        sl = QHBoxLayout(sec); sl.setContentsMargins(16,0,14,0)
        sl.addWidget(QLabel("Ket qua dang nhap", styleSheet="color:#8da2cc;font-size:12px;background:transparent;"))
        right_v.addWidget(sec)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Username", "Password", "Session ID", "Status", "Status Account"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        for col, w in [(2, 150), (3, 110), (4, 160)]:
            self.table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Fixed)
            self.table.setColumnWidth(col, w)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True); self.table.setShowGrid(False)
        right_v.addWidget(self.table, 1)

        log_hdr = QWidget(); log_hdr.setFixedHeight(30)
        log_hdr.setStyleSheet("background:#0f131b; border-top:1px solid #1a2030; border-bottom:1px solid #1a2030;")
        lhh = QHBoxLayout(log_hdr); lhh.setContentsMargins(14,0,14,0)
        lhh.addWidget(QLabel("LOG", styleSheet="color:#5a7090;font-size:10px;font-weight:bold;background:transparent;"))
        lhh.addStretch()
        cb = QPushButton("Xoa"); cb.setObjectName("btn_clear"); cb.setFixedHeight(20)
        cb.clicked.connect(self._clear_log); lhh.addWidget(cb)
        right_v.addWidget(log_hdr)

        self.log_box = QTextEdit(); self.log_box.setObjectName("log_box")
        self.log_box.setReadOnly(True); self.log_box.setFixedHeight(110)
        right_v.addWidget(self.log_box)

        body.addWidget(right_w, 1)
        root_v.addLayout(body, 1)

    def _on_start(self):
        if core_login._running: return
        self.btn_start.setEnabled(False); self.btn_stop.setEnabled(True)
        core_login.trigger_run(_loop)

    def _on_log(self, msg): self.log_box.append(msg)
    def _clear_log(self): self.log_box.clear()
    def _tick(self): self._refresh_sessions()

    def _on_refresh(self):
        self._refresh_sessions(); self._refresh_table(); self._refresh_stats()
        if not core_login._running:
            self.btn_start.setEnabled(True); self.btn_stop.setEnabled(False)

    def _refresh_sessions(self):
        self.sess_list.clear()
        for sid, info in core_login.sessions.items():
            elapsed = int(time.time() - info.get("connected_at", time.time()))
            m, s = divmod(elapsed, 60)
            item = QListWidgetItem(f"{str(sid)[:22]}...\n{info.get('status','-')}  —  {m:02d}:{s:02d}")
            self.sess_list.addItem(item)
        self.count_badge.setText(f"{len(core_login.sessions)} ket noi")

    def _get_ban_status(self, e):
        """Trả về (text, bg_hex, fg_hex) dựa trên ban restrictions trong userinfo."""
        # Chưa fetch userinfo → hiện “-”
        if "userinfo" not in e:
            return "-", "#111620", "#4a6080"
        ui = e.get("userinfo", {})
    
        # userinfo là dict rỗng → fetch lỗi hoặc không có data
        if not isinstance(ui, dict) or not ui:
            return "No data", "#1a1a20", "#6060a0"
        restrictions = ui.get("ban", {}).get("restrictions", None)
        print(f"restrictions: {restrictions}")
        # restrictions là None hoặc list rỗng → sạch
        if not restrictions:
            return "OK", "#0a1a0f", "#4ae98c"
        # Kiểm tra PERMANENT_BAN do script
        for r in restrictions:
            if r.get("type") == "PERMANENT_BAN":
                reason = r.get("reason", "")
                if "SCRIPTING" in reason:
                    return "PERM BAN (script)", "#2a0a0a", "#ff4444"
                return "PERMANENT BAN", "#2a1a0a", "#ff8844"
        # Có ban nhưng không phải PERMANENT
        return "BANNED", "#1a1a0a", "#ffcc44"

    def _refresh_table(self):
        self.table.setRowCount(len(core_login.acct_log))
        for row, e in enumerate(core_login.acct_log):
            status = e.get("status", "Cho...")
            bg_hex, fg_hex = STATUS_COLORS.get(status, ("#0f131b", "#8da0c0"))
            bg, fg = QColor(bg_hex), QColor(fg_hex)
            for col, val, align in [
                (0, e.get("username",""), None),
                (1, e.get("password",""), None),
                (2, str(e["session_id"])[:22] + "...", None),
                (3, status, CENTER),
            ]:
                self.table.setItem(row, col, make_cell(val, bg, fg, align or (Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)))
            # ── Cột Status Account (ban check) ──
            ban_text, ban_bg, ban_fg = self._get_ban_status(e)
            self.table.setItem(row, 4, make_cell(ban_text, QColor(ban_bg), QColor(ban_fg), CENTER))

    def _refresh_stats(self):
        log = core_login.acct_log
        self.s_conn.setText(str(len(core_login.sessions)))
        self.s_running.setText(str(sum(1 for e in log if "chay" in e.get("status","").lower())))
        self.s_done.setText(str(sum(1 for e in log if "Hoan" in e.get("status",""))))
        self.s_mfa.setText(str(sum(1 for e in log if "2FA" in e.get("status",""))))
        self.s_error.setText(str(sum(1 for e in log if "Loi" in e.get("status",""))))

    def _export_excel(self):
        if not core_login.acct_log:
            QMessageBox.information(self, "Thong bao", "Chua co du lieu de xuat!")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Luu file", "login_results.xlsx", "Excel Files (*.xlsx)")
        if not path: return
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Login Results"
        ws.append(["Username", "Password", "Status", "Note", "Cookies"])
        for e in core_login.acct_log:
            ws.append([e.get('username',''), e.get('password',''), e.get('status',''), e.get('note',''), e.get('cookies','')])
        wb.save(path)
        QMessageBox.information(self, "Thanh cong", f"Da xuat {len(core_login.acct_log)} dong ra:\n{path}")


# ══════════════════════════════════════════════════════════════════════════════
# Tab 2: CHANGE PASSWORD
# ══════════════════════════════════════════════════════════════════════════════
class ChangepassTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build()
        core_changepass.bridge.log_signal.connect(self._on_log)
        core_changepass.bridge.refresh_signal.connect(self._on_refresh)

        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(1000)

    def _build(self):
        root_v = QVBoxLayout(self)
        root_v.setContentsMargins(0, 0, 0, 0); root_v.setSpacing(0)

        tb = QWidget(); tb.setObjectName("toolbar"); tb.setFixedHeight(50)
        h = QHBoxLayout(tb); h.setContentsMargins(16, 0, 16, 0); h.setSpacing(8)
        t = QLabel("Doi mat khau — Riot Account")
        t.setStyleSheet("color:#8da2cc; font-size:12px; font-weight:bold; background:transparent;")
        h.addWidget(t); h.addStretch()

        self.btn_stop  = QPushButton("■  STOP")
        self.btn_stop.setObjectName("btn_stop"); self.btn_stop.setFixedHeight(32)
        self.btn_stop.setEnabled(False); self.btn_stop.clicked.connect(core_changepass.trigger_stop)

        self.btn_export = QPushButton("XUAT EXCEL")
        self.btn_export.setObjectName("btn_export"); self.btn_export.setFixedHeight(32)
        self.btn_export.clicked.connect(self._export_excel)

        self.btn_start = QPushButton("▶  BAT DAU")
        self.btn_start.setObjectName("btn_start_changepass"); self.btn_start.setFixedHeight(32)
        self.btn_start.clicked.connect(self._on_start)

        for w in [self.btn_stop, self.btn_export, self.btn_start]:
            h.addWidget(w)
        root_v.addWidget(tb)

        body = QHBoxLayout()
        body.setContentsMargins(0, 0, 0, 0); body.setSpacing(0)

        left = QWidget(); left.setObjectName("left_panel"); left.setFixedWidth(230)
        lv = QVBoxLayout(left); lv.setContentsMargins(0,0,0,0); lv.setSpacing(0)
        lhdr = QWidget(); lhdr.setObjectName("left_header"); lhdr.setFixedHeight(46)
        lh = QHBoxLayout(lhdr); lh.setContentsMargins(12, 0, 12, 0)
        lh.addWidget(QLabel("Session ID", styleSheet="color:white;font-weight:bold;background:transparent;"))
        lh.addStretch()
        self.count_badge = QLabel("0 ket noi"); self.count_badge.setObjectName("count_badge")
        lh.addWidget(self.count_badge)
        lv.addWidget(lhdr)
        self.sess_list = QListWidget(); lv.addWidget(self.sess_list, 1)
        body.addWidget(left)

        right_w = QWidget(); right_v = QVBoxLayout(right_w)
        right_v.setContentsMargins(0,0,0,0); right_v.setSpacing(0)

        cards_w = QWidget()
        cards_w.setStyleSheet("background:#0b0e14; border-bottom:1px solid #262d38;")
        cards_h = QHBoxLayout(cards_w); cards_h.setContentsMargins(0,0,0,0); cards_h.setSpacing(0)
        self.s_conn    = make_stat_card(cards_h, "KET NOI",    "0", "#4282ff", last=False)
        self.s_running = make_stat_card(cards_h, "DANG CHAY",  "0", "#fdb45a", last=False)
        self.s_done    = make_stat_card(cards_h, "HOAN THANH", "0", "#4ae98c", last=False)
        self.s_error   = make_stat_card(cards_h, "LOI",        "0", "#ff7070", last=True)
        cards_w.setFixedHeight(90); right_v.addWidget(cards_w)

        sec = QWidget(); sec.setFixedHeight(34)
        sec.setStyleSheet("background:#121a24; border-bottom:1px solid #263040;")
        sl = QHBoxLayout(sec); sl.setContentsMargins(16,0,14,0)
        sl.addWidget(QLabel("Danh sach doi mat khau", styleSheet="color:#8da2cc;font-size:12px;background:transparent;"))
        right_v.addWidget(sec)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Username", "Mat khau cu", "Mat khau moi", "Session ID", "Status"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        for col, w in [(3, 150), (4, 100)]:
            self.table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Fixed)
            self.table.setColumnWidth(col, w)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True); self.table.setShowGrid(False)
        right_v.addWidget(self.table, 1)

        log_hdr = QWidget(); log_hdr.setFixedHeight(30)
        log_hdr.setStyleSheet("background:#0f131b; border-top:1px solid #1a2030; border-bottom:1px solid #1a2030;")
        lhh = QHBoxLayout(log_hdr); lhh.setContentsMargins(14,0,14,0)
        lhh.addWidget(QLabel("LOG", styleSheet="color:#5a7090;font-size:10px;font-weight:bold;background:transparent;"))
        lhh.addStretch()
        cb = QPushButton("Xoa"); cb.setObjectName("btn_clear"); cb.setFixedHeight(20)
        cb.clicked.connect(self._clear_log); lhh.addWidget(cb)
        right_v.addWidget(log_hdr)

        self.log_box = QTextEdit(); self.log_box.setObjectName("log_box")
        self.log_box.setReadOnly(True); self.log_box.setFixedHeight(110)
        right_v.addWidget(self.log_box)

        body.addWidget(right_w, 1)
        root_v.addLayout(body, 1)

    def _on_start(self):
        if core_changepass._running: return
        self.btn_start.setEnabled(False); self.btn_stop.setEnabled(True)
        core_changepass.trigger_run(_loop)

    def _on_log(self, msg): self.log_box.append(msg)
    def _clear_log(self): self.log_box.clear()
    def _tick(self): self._refresh_sessions()

    def _on_refresh(self):
        self._refresh_sessions(); self._refresh_table(); self._refresh_stats()
        if not core_changepass._running:
            self.btn_start.setEnabled(True); self.btn_stop.setEnabled(False)

    def _refresh_sessions(self):
        self.sess_list.clear()
        for sid, info in core_changepass.sessions.items():
            elapsed = int(time.time() - info.get("connected_at", time.time()))
            m, s = divmod(elapsed, 60)
            item = QListWidgetItem(f"{str(sid)[:22]}...\n{info.get('status','-')}  —  {m:02d}:{s:02d}")
            self.sess_list.addItem(item)
        self.count_badge.setText(f"{len(core_changepass.sessions)} ket noi")

    def _refresh_table(self):
        self.table.setRowCount(len(core_changepass.acct_log))
        for row, e in enumerate(core_changepass.acct_log):
            status = e.get("status", "Cho...")
            bg_hex, fg_hex = STATUS_COLORS.get(status, ("#0f131b", "#8da0c0"))
            bg, fg = QColor(bg_hex), QColor(fg_hex)
            for col, val, align in [
                (0, e.get("username",""), None),
                (1, e.get("old_password",""), None),
                (2, e.get("new_password",""), CENTER),
                (3, str(e["session_id"])[:22] + "...", None),
                (4, status, CENTER),
            ]:
                self.table.setItem(row, col, make_cell(val, bg, fg, align or (Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)))

    def _refresh_stats(self):
        log = core_changepass.acct_log
        self.s_conn.setText(str(len(core_changepass.sessions)))
        self.s_running.setText(str(sum(1 for e in log if "chay" in e.get("status","").lower())))
        self.s_done.setText(str(sum(1 for e in log if "Hoan" in e.get("status",""))))
        self.s_error.setText(str(sum(1 for e in log if "Loi" in e.get("status",""))))

    def _export_excel(self):
        if not core_changepass.acct_log:
            QMessageBox.information(self, "Thong bao", "Chua co du lieu de xuat!")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Luu file", "changed_accounts.xlsx", "Excel Files (*.xlsx)")
        if not path: return
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Change Password Results"
        ws.append(["Username", "New Password", "Status"])
        for e in core_changepass.acct_log:
            ws.append([e.get('username',''), e.get('new_password',''), e.get('status','')])
        wb.save(path)
        QMessageBox.information(self, "Thanh cong", f"Da xuat {len(core_changepass.acct_log)} dong ra:\n{path}")


# ══════════════════════════════════════════════════════════════════════════════
# Tab 3: REGISTER
# ══════════════════════════════════════════════════════════════════════════════
class RegisterTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build()
        core_register.bridge.log_signal.connect(self._on_log)
        core_register.bridge.refresh_signal.connect(self._on_refresh)

        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(1000)

    def _build(self):
        root_v = QVBoxLayout(self)
        root_v.setContentsMargins(0, 0, 0, 0); root_v.setSpacing(0)

        tb = QWidget(); tb.setObjectName("toolbar"); tb.setFixedHeight(50)
        h = QHBoxLayout(tb); h.setContentsMargins(16, 0, 16, 0); h.setSpacing(8)
        t = QLabel("Tao tai khoan — League of Legends")
        t.setStyleSheet("color:#8da2cc; font-size:12px; font-weight:bold; background:transparent;")
        h.addWidget(t); h.addStretch()

        self.btn_stop  = QPushButton("■  STOP")
        self.btn_stop.setObjectName("btn_stop"); self.btn_stop.setFixedHeight(32)
        self.btn_stop.setEnabled(False); self.btn_stop.clicked.connect(core_register.trigger_stop)

        self.btn_export = QPushButton("XUAT EXCEL")
        self.btn_export.setObjectName("btn_export"); self.btn_export.setFixedHeight(32)
        self.btn_export.clicked.connect(self._export_excel)

        self.btn_start = QPushButton("▶  BAT DAU")
        self.btn_start.setObjectName("btn_start_register"); self.btn_start.setFixedHeight(32)
        self.btn_start.clicked.connect(self._on_start)

        for w in [self.btn_stop, self.btn_export, self.btn_start]:
            h.addWidget(w)
        root_v.addWidget(tb)

        body = QHBoxLayout()
        body.setContentsMargins(0, 0, 0, 0); body.setSpacing(0)

        left = QWidget(); left.setObjectName("left_panel"); left.setFixedWidth(230)
        lv = QVBoxLayout(left); lv.setContentsMargins(0,0,0,0); lv.setSpacing(0)
        lhdr = QWidget(); lhdr.setObjectName("left_header"); lhdr.setFixedHeight(46)
        lh = QHBoxLayout(lhdr); lh.setContentsMargins(12, 0, 12, 0)
        lh.addWidget(QLabel("Session ID", styleSheet="color:white;font-weight:bold;background:transparent;"))
        lh.addStretch()
        self.count_badge = QLabel("0 ket noi"); self.count_badge.setObjectName("count_badge")
        lh.addWidget(self.count_badge)
        lv.addWidget(lhdr)
        self.sess_list = QListWidget(); lv.addWidget(self.sess_list, 1)
        body.addWidget(left)

        right_w = QWidget(); right_v = QVBoxLayout(right_w)
        right_v.setContentsMargins(0,0,0,0); right_v.setSpacing(0)

        cards_w = QWidget()
        cards_w.setStyleSheet("background:#0b0e14; border-bottom:1px solid #262d38;")
        cards_h = QHBoxLayout(cards_w); cards_h.setContentsMargins(0,0,0,0); cards_h.setSpacing(0)
        self.s_conn    = make_stat_card(cards_h, "KET NOI",    "0", "#4282ff", last=False)
        self.s_running = make_stat_card(cards_h, "DANG CHAY",  "0", "#fdb45a", last=False)
        self.s_done    = make_stat_card(cards_h, "HOAN THANH", "0", "#4ae98c", last=True)
        cards_w.setFixedHeight(90); right_v.addWidget(cards_w)

        sec = QWidget(); sec.setFixedHeight(34)
        sec.setStyleSheet("background:#121a24; border-bottom:1px solid #263040;")
        sl = QHBoxLayout(sec); sl.setContentsMargins(16,0,14,0)
        sl.addWidget(QLabel("Danh sach tai khoan dang ky", styleSheet="color:#8da2cc;font-size:12px;background:transparent;"))
        right_v.addWidget(sec)

        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["User Name", "Email", "Password", "Token", "Session ID", "Status"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        for col, w in [(2, 110), (3, 180), (4, 150), (5, 100)]:
            self.table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Fixed)
            self.table.setColumnWidth(col, w)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True); self.table.setShowGrid(False)
        right_v.addWidget(self.table, 1)

        log_hdr = QWidget(); log_hdr.setFixedHeight(30)
        log_hdr.setStyleSheet("background:#0f131b; border-top:1px solid #1a2030; border-bottom:1px solid #1a2030;")
        lhh = QHBoxLayout(log_hdr); lhh.setContentsMargins(14,0,14,0)
        lhh.addWidget(QLabel("LOG", styleSheet="color:#5a7090;font-size:10px;font-weight:bold;background:transparent;"))
        lhh.addStretch()
        cb = QPushButton("Xoa"); cb.setObjectName("btn_clear"); cb.setFixedHeight(20)
        cb.clicked.connect(self._clear_log); lhh.addWidget(cb)
        right_v.addWidget(log_hdr)

        self.log_box = QTextEdit(); self.log_box.setObjectName("log_box")
        self.log_box.setReadOnly(True); self.log_box.setFixedHeight(110)
        right_v.addWidget(self.log_box)

        body.addWidget(right_w, 1)
        root_v.addLayout(body, 1)

    def _on_start(self):
        if core_register._running: return
        self.btn_start.setEnabled(False); self.btn_stop.setEnabled(True)
        core_register.trigger_run(_loop)

    def _on_log(self, msg): self.log_box.append(msg)
    def _clear_log(self): self.log_box.clear()
    def _tick(self): self._refresh_sessions()

    def _on_refresh(self):
        self._refresh_sessions(); self._refresh_table(); self._refresh_stats()
        if not core_register._running:
            self.btn_start.setEnabled(True); self.btn_stop.setEnabled(False)

    def _refresh_sessions(self):
        self.sess_list.clear()
        for sid, info in core_register.sessions.items():
            elapsed = int(time.time() - info.get("connected_at", time.time()))
            m, s = divmod(elapsed, 60)
            item = QListWidgetItem(f"{str(sid)[:22]}...\n{info.get('status','-')}  —  {m:02d}:{s:02d}")
            self.sess_list.addItem(item)
        self.count_badge.setText(f"{len(core_register.sessions)} ket noi")

    def _refresh_table(self):
        self.table.setRowCount(len(core_register.acct_log))
        for row, e in enumerate(core_register.acct_log):
            status = e.get("status", "Cho...")
            bg_hex, fg_hex = STATUS_COLORS.get(status, ("#0f131b", "#8da0c0"))
            bg, fg = QColor(bg_hex), QColor(fg_hex)
            raw_token   = e.get("token", "")
            short_token = (raw_token[:30] + "...") if len(raw_token) > 30 else raw_token
            for col, val, align in [
                (0, e.get("username",""), None),
                (1, e.get("email",""), None),
                (2, e.get("password",""), CENTER),
                (3, short_token, None),
                (4, str(e["session_id"])[:22] + "...", None),
                (5, status, CENTER),
            ]:
                self.table.setItem(row, col, make_cell(val, bg, fg, align or (Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)))

    def _refresh_stats(self):
        log = core_register.acct_log
        self.s_conn.setText(str(len(core_register.sessions)))
        self.s_running.setText(str(sum(1 for e in log if "chay" in e.get("status","").lower())))
        self.s_done.setText(str(sum(1 for e in log if "Hoan" in e.get("status",""))))

    def _export_excel(self):
        if not core_register.acct_log:
            QMessageBox.information(self, "Thong bao", "Chua co du lieu de xuat!")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Luu file", "accounts.xlsx", "Excel Files (*.xlsx)")
        if not path: return
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Register Results"
        ws.append(["Username", "Email", "Password", "Token"])
        for e in core_register.acct_log:
            ws.append([e.get('username',''), e.get('email',''), e.get('password',''), e.get('token','')])
        wb.save(path)
        QMessageBox.information(self, "Thanh cong", f"Da xuat {len(core_register.acct_log)} tai khoan ra:\n{path}")


# ══════════════════════════════════════════════════════════════════════════════
# Main Window
# ══════════════════════════════════════════════════════════════════════════════
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Hieusoft Riot")
        self.setMinimumSize(1150, 720)

        root = QWidget(); root.setObjectName("root")
        self.setCentralWidget(root)
        root_v = QVBoxLayout(root)
        root_v.setContentsMargins(0, 0, 0, 0); root_v.setSpacing(0)

        # ── Toolbar ──
        tb = QWidget(); tb.setObjectName("toolbar"); tb.setFixedHeight(56)
        h = QHBoxLayout(tb); h.setContentsMargins(18, 0, 18, 0); h.setSpacing(0)

        logo_box = QWidget(); logo_box.setFixedSize(96, 36)
        logo_box.setStyleSheet("background: qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1a88b8,stop:1 #1d9e5d); border:1px solid #2eb8d8;")
        logo_inner = QHBoxLayout(logo_box); logo_inner.setContentsMargins(0,0,0,0)
        logo_lbl = QLabel("HieuSoft")
        logo_lbl.setStyleSheet("color:white; font-size:13px; font-weight:bold; background:transparent;")
        logo_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_inner.addWidget(logo_lbl)
        h.addWidget(logo_box)

        title = QLabel("  Tool-RIOT")
        title.setStyleSheet("color:white; font-size:15px; font-weight:bold; background:transparent;")
        h.addWidget(title)
        h.addStretch()

        self.ws_label = QLabel("▸ WS: khoi dong...")
        self.ws_label.setStyleSheet("color:#4a6080; font-size:11px; background:transparent;")
        h.addWidget(self.ws_label)

        root_v.addWidget(tb)

        # ── Tabs ──
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.TabPosition.North)
        self.tab_login      = LoginTab()
        self.tab_changepass = ChangepassTab()
        self.tab_register   = RegisterTab()
        self.tabs.addTab(self.tab_login,      "  Login  ")
        self.tabs.addTab(self.tab_changepass, "  Change Password  ")
        self.tabs.addTab(self.tab_register,   "  Register  ")
        root_v.addWidget(self.tabs, 1)

        # ── Status bar ──
        self.status_bar = QLabel("San sang  |  WebSocket: ws://127.0.0.1:8000")
        self.status_bar.setObjectName("status_bar"); self.status_bar.setFixedHeight(26)
        root_v.addWidget(self.status_bar)

        # Kết nối signal từ WS thread → main thread
        ws_bridge.server_ready_signal.connect(self._on_server_ready)

    def _on_server_ready(self):
        self.ws_label.setStyleSheet("color:#4ae98c; font-size:11px; background:transparent;")
        self.ws_label.setText("● WS: ws://127.0.0.1:8000")
        self.status_bar.setText("Server san sang: ws://127.0.0.1:8000  ✓")


# ══════════════════════════════════════════════════════════════════════════════
# WebSocket server chung — route session đến đúng module
# ══════════════════════════════════════════════════════════════════════════════

# Bridge để thông báo WS server đã sẵn sàng sang Qt main thread
class _WsBridge(QObject):
    server_ready_signal = pyqtSignal()

ws_bridge = _WsBridge()

MODULES = {
    "login":      core_login,
    "changepass": core_changepass,
    "register":   core_register,
}

# Map session_id → tên module (được set khi extension gửi register kèm "mode")
_session_module = {}   # {session_id: "login" | "changepass" | "register"}

_main_window = None   # set từ run.py

import websockets

async def ws_handler(websocket):
    session_id = None
    mod = None
    try:
        async for raw in websocket:
            try:
                msg = json.loads(raw)
            except Exception as e:
                print(f"[WS] Parse error: {e}")
                continue

            t = msg.get("type")
            if t == "register":
                session_id = msg.get("sessionId")
                if not session_id:
                    continue

                # Extension gửi "mode": "login" | "changepass" | "register"
                # Nếu không có → mặc định "login"
                mode = msg.get("mode", "login").lower()
                if mode not in MODULES:
                    mode = "login"

                mod = MODULES[mode]
                _session_module[session_id] = mode

                mod.sessions[session_id] = {
                    "ws": websocket, "info": msg,
                    "pending": {}, "status": "Ket noi",
                    "connected_at": time.time()
                }
                print(f"[WS] ++ Session [{str(session_id)[:8]}] mode={mode} ({sum(len(m.sessions) for m in MODULES.values())} total)")
                mod.bridge.refresh_signal.emit()

            elif t == "result":
                # Route về đúng module
                rid = msg.get("requestId")
                sid = msg.get("sessionId") or session_id
                if sid is None:
                    continue
                m_name = _session_module.get(sid)
                if not m_name:
                    continue
                m = MODULES[m_name]
                if sid in m.sessions:
                    p = m.sessions[sid]["pending"]
                    if rid and rid in p and not p[rid].done():
                        p[rid].set_result(msg)

    except websockets.exceptions.ConnectionClosed:
        pass
    except Exception as e:
        print(f"[WS] Handler error: {e}")
    finally:
        if session_id and mod and session_id in mod.sessions:
            mod.sessions.pop(session_id, None)
            _session_module.pop(session_id, None)
            print(f"[WS] -- Session ngat: {str(session_id)[:8]}")
            mod.bridge.refresh_signal.emit()


async def _ws_server_coro():
    async with websockets.serve(ws_handler, "127.0.0.1", 8000):
        print("[WS] Server san sang: ws://127.0.0.1:8000")
        ws_bridge.server_ready_signal.emit()   # thread-safe, Qt sẽ gọi về main thread
        await asyncio.Future()   # chạy mãi


def start_asyncio_loop():
    global _loop
    _loop = asyncio.new_event_loop()
    asyncio.set_event_loop(_loop)
    # Chia sẻ loop cho các core module
    core_login._loop      = _loop
    core_changepass._loop = _loop
    core_register._loop   = _loop
    _loop.run_until_complete(_ws_server_coro())
